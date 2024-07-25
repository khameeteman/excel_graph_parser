from io import BytesIO
from pathlib import Path
from typing import List, Union, Optional

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook
from viktor import UserError, UserMessage, File
from viktor.errors import InputViolation
from viktor.external.spreadsheet import SpreadsheetCalculationInput, SpreadsheetCalculation

import plotly.graph_objects as go

ALLOWED_FIGURE_TYPES = ["lineChart", "scatterChart", "barChart", "pieChart"]


class ExcelImageParser:
    def __init__(self, excel_file_path: Union[Path, str] = None, params: dict = None, from_app: bool = False,
                 *, spreadsheet_calculation: SpreadsheetCalculation = None):

        if spreadsheet_calculation is not None:
            file = spreadsheet_calculation._file
            if isinstance(file, File):
                with file.open_binary() as r:
                    self.workbook = load_workbook(filename=r, data_only=True)
            elif isinstance(file, BytesIO):
                self.workbook = load_workbook(filename=file, data_only=True)
            else:
                raise NotImplementedError
        else:
            self.workbook = load_workbook(filename=excel_file_path, data_only=True)

        self.excel_file_path = excel_file_path
        self.params = params
        self.from_app = from_app
        self._spreadsheet_calculation = spreadsheet_calculation

        # Add exit point name to dataframe
        sheet_names = self.workbook.sheetnames

        # Gather charts by looping through sheets
        self.charts = []
        self._charts_map = {}

        untitled_index = 1
        for sheet_name in sheet_names:
            sheet = self.workbook[sheet_name]
            for chart in sheet._charts:  # could be empty
                if chart.title:
                    chart_title = ''.join([title_element.t for title_element in chart.title.tx.rich.p[0].r])
                else:
                    chart_title = f"Untitled {untitled_index}"
                    untitled_index += 1

                self._charts_map[chart_title] = chart
                self.charts.append(chart)

    def validate_sheet_names(self):
        """Validate that the input sheet and output sheets are present"""
        wb = self.workbook
        if not all(sheetname in wb for sheetname in ["viktor-input-sheet", "viktor-output-sheet"]):
            raise UserError(
                "The sheet names are not correctly formatted.",
                input_violations=[
                    InputViolation(message="Please check the sheet and follow the documentation", fields=["excel_file"])
                ],
            )

    def get_input_cells(self) -> List[dict]:
        """Gets inputs from the excel file as a dict"""
        wb = self.workbook
        ws_input = wb["viktor-input-sheet"]
        inputs = []
        for index, row in enumerate(ws_input.iter_rows(min_row=2, max_col=4)):
            if row[0].value:
                inputs.append(
                    {
                        "name": row[0].value,
                        "unit": row[1].value if row[1].value else "",
                        "description": row[2].value,
                        "default": row[3].value,
                        "key": f"input_{index}",
                    }
                )
        return inputs

    def get_evaluated_spreadsheet(self):
        """Evaluate spreadsheet so the version with the updated inputs and outputs is available"""
        inputs = []

        if self._spreadsheet_calculation is not None:
            spreadsheet = self._spreadsheet_calculation
        else:
            input_cells = self.get_input_cells()

            # Check whether the user wrongfully adjusted the inputs table
            if not self.from_app:
                if len(input_cells) != len(self.params.preview_step.fields_table):
                    raise UserError(
                        "Please do not add or delete rows from the input table, go back to the previous step and re-process"
                        " the uploaded file"
                    )

            # Load spreadsheet with correct inputs
            if not self.from_app:
                for (row, input_cell) in zip(self.params.preview_step.fields_table, input_cells):
                    field_name = input_cell["name"]
                    value = row["values"]
                    inputs.append(SpreadsheetCalculationInput(field_name, value))

                spreadsheet = SpreadsheetCalculation(self.params.upload_step.excel_file.file, inputs)

            else:
                for input_cell in input_cells:
                    field_name = input_cell["name"]
                    value = self.params[input_cell["key"]]
                    inputs.append(SpreadsheetCalculationInput(field_name, value))

                spreadsheet = SpreadsheetCalculation.from_path(self.excel_file_path, inputs)

        result = spreadsheet.evaluate(include_filled_file=True)
        evaluated_workbook = load_workbook(BytesIO(result.file_content), data_only=True)

        return evaluated_workbook, result

    def get_outputs(self) -> List[dict]:
        """Gets outputs from the excel file as a dict (will return empty if no outputs are present in sheet)"""
        wb, result = self.get_evaluated_spreadsheet()
        values = result.values
        ws_output = wb["viktor-output-sheet"]
        outputs = []
        if not values:
            return outputs
        for index, row in enumerate(ws_output.iter_rows(min_row=2, max_col=4)):
            if row[0].value:
                outputs.append(
                    {
                        "name": row[0].value,
                        "unit": row[1].value if row[1].value else "",
                        "description": row[2].value,
                        "key": f"output_{index}",
                        "value": values[row[0].value],
                        "type": str(type(values[row[0].value]))
                    }
                )
        return outputs

    def get_figures_from_excel_file(self) -> list:
        """Gets figures from the excel file as a list"""
        wb, _ = self.get_evaluated_spreadsheet()

        figures = []
        for chart_title in self._charts_map:
            chart_data = self._parse_chart_data(chart_title, wb)
            if chart_data is None:
                continue
            chart_data['fig'] = self._create_plotly_figure(chart_data)
            figures.append(chart_data)

        wb.close()
        return figures

    def get_plotly_figure_by_title(self, title: str) -> go.Figure:
        """Gets plotly figure by title"""
        wb, _ = self.get_evaluated_spreadsheet()
        chart_data = self._parse_chart_data(title, wb)
        wb.close()

        if chart_data is None:
            raise UserError(f"No figure found with title: {title}")

        return self._create_plotly_figure(chart_data)

    def get_figure_titles(self) -> List[dict]:
        """Generate dict with all the names of each figure to include in app template"""
        figure_list = []

        for chart_title, chart in self._charts_map.items():
            clean_name = [s.lower() for s in chart_title.replace(" ", "_") if s.isalnum() or s == "_"]
            figure_name = "".join(clean_name)
            figure_type = chart.tagname
            figure_list.append(
                {
                    "name": chart_title,
                    "concat_name": figure_name,
                    "type": figure_type,
                }
            )

        return figure_list

    def _parse_chart_data(self, chart_title: str, wb: Workbook) -> Optional[dict]:
        """Extracts chart data from the provided workbook"""
        chart = self._charts_map[chart_title]

        # Get the general chart elements
        series = []
        chart_type = chart.tagname
        if chart_type not in ALLOWED_FIGURE_TYPES:
            UserMessage.warning(
                f"Chart titled '{chart_title}' is not of one of the allowed types and can not be visualised"
            )
            return None

        x_axis_title, y_axis_title = None, None
        if chart_type != "pieChart":
            if chart.x_axis.title:
                x_axis_title = chart.x_axis.title.tx.rich.p[0].r[-1].t
            if chart.y_axis.title:
                y_axis_title = chart.y_axis.title.tx.rich.p[0].r[-1].t

        # Get series data
        input_cat_range = ""
        input_cat_format = None
        for i, series in enumerate(chart.series):
            if chart_type == "scatterChart":
                if series.xVal:
                    if series.xVal.strRef:
                        input_cat_range = series.xVal.strRef.f
                    elif series.xVal.numRef:
                        input_cat_range = series.xVal.numRef.f
                        input_cat_format = series.xVal.numRef.numCache.formatCode

                input_val_range = series.yVal.numRef.f
                input_val_format = series.yVal.numRef.numCache.formatCode

            else:
                if series.cat:
                    # if no category data in the sequence, use the one that was set for the previous sequence
                    if series.cat.strRef:
                        input_cat_range = series.cat.strRef.f
                    elif series.cat.numRef:
                        input_cat_range = series.cat.numRef.f
                        input_cat_format = series.cat.numRef.numCache.formatCode

                input_val_range = series.val.numRef.f
                input_val_format = series.val.numRef.numCache.formatCode

            input_cat_format = None if input_cat_format == "General" else input_cat_format
            input_val_format = None if input_val_format == "General" else input_val_format

            # category_axis_data
            chart_sheet_name = (
                input_cat_range
                .replace("(", "")
                .replace(")", "")
                .replace("'", "")
                .split(sep="!")[0]
            )
            chart_cat_range = (
                input_cat_range
                .replace('(', '')
                .replace(')', '')
                .replace("'", "")
                .replace(f"{chart_sheet_name}!", "")
                .replace('$', "")
            )
            chart_cat_range = chart_cat_range.split(",")[0] + ":" + chart_cat_range.split(",")[-1] if "," in chart_cat_range else chart_cat_range
            cat_data = []
            for element in wb[chart_sheet_name][chart_cat_range]:
                for sub_element in element:
                    if type(sub_element) == Cell:
                        cat_data.append(sub_element.value)

            # value_axis_data
            chart_sheet_name = (
                input_val_range
                .replace('(', '')
                .replace(')', '')
                .replace("'", "")
                .split(sep="!")[0]
            )
            chart_val_range = (
                input_val_range
                .replace('(', '')
                .replace(')', '')
                .replace("'", "")
                .replace(f"{chart_sheet_name}!", "")
            )
            chart_val_range = chart_val_range.split(",")[0] + ":" + chart_val_range.split(",")[-1] if "," in chart_val_range else chart_val_range
            val_data = []
            for element in wb[chart_sheet_name][chart_val_range]:
                for sub_element in element:
                    if type(sub_element) == Cell:
                        val_data.append(sub_element.value)

            series_name = series.tx.v if series.tx else None
            ser = {
                "category_axis_data": cat_data,
                "value_axis_data": val_data,
                "category_value_format": input_cat_format,
                "values_value_format": input_val_format,
                "series_name": series_name if series_name else None
            }
            series.append(ser)

        chart_data = {
            "chart_title": chart_title,
            "chart_type": chart_type,
            "x_axis_title": x_axis_title,
            "y_axis_title": y_axis_title,
            "series": series,
        }

        return chart_data

    @staticmethod
    def _create_plotly_figure(chart_data: dict) -> go.Figure:
        """Creates plotly figure based on the extracted chart data"""
        fig = go.Figure()
        if chart_data["chart_type"] == "lineChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Scatter(x=ser["category_axis_data"], y=ser["value_axis_data"], mode='lines', name=ser["series_name"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )
        if chart_data["chart_type"] == "barChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Bar(x=ser["category_axis_data"], y=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )
        if chart_data["chart_type"] == "pieChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Pie(labels=ser["category_axis_data"], values=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
            )
        if chart_data["chart_type"] == "scatterChart":
            for ser in chart_data["series"]:
                fig.add_trace(go.Scatter(x=ser["category_axis_data"], y=ser["value_axis_data"]))
            fig.update_layout(
                title_text=chart_data["chart_title"],
                xaxis_title=chart_data["x_axis_title"],
                yaxis_title=chart_data["y_axis_title"],
                yaxis_tickformat=chart_data["series"][0]["values_value_format"],
                xaxis_tickformat=chart_data["series"][0]["category_value_format"],
            )

        return fig
